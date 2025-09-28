
"""
excel_playbooks_append.py
Append a "Playbooks" sheet to an existing Excel file using openpyxl.
"""

from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def append_playbooks_sheet(xlsx_path: str, playbooks: List[Dict[str, Any]]) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.create_sheet("Playbooks")
    headers = ["ID", "Title", "Rationale", "Steps", "Expected Outcome"]
    ws.append(headers)
    for pb in playbooks:
        steps = pb.get("steps", [])
        ws.append([
            pb.get("id", ""),
            pb.get("title", ""),
            pb.get("rationale", ""),
            "\n".join(f"- {s}" for s in steps),
            pb.get("expected_outcome", ""),
        ])
    # set column widths
    widths = [10, 30, 40, 60, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(xlsx_path)
