import os
import sys
import pandas as pd
from io import BytesIO
from fastapi.testclient import TestClient
import openpyxl

# ทำให้ import main/pdf_summary/utils จากโฟลเดอร์ budget_plus ได้เมื่อรันจาก budget_plus/
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))    # .../budget_plus/tests
PROJECT_ROOT = os.path.dirname(CURRENT_DIR)                 # .../budget_plus
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from budget_plus.main import app
from budget_plus.pdf_summary import generate_pdf_with_chart
from budget_plus.utils.number_format_utils import format_number

client = TestClient(app)


def _make_excel(df: pd.DataFrame) -> BytesIO:
    buffer = BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)
    return buffer


def test_multi_version_variance():
    """API /analyze ต้องคืนค่าเป็น string ที่ format แล้ว"""
    df = pd.DataFrame({
        "Version": ["V1", "V2"],
        "Scenario": ["Base", "What-if"],
        "Cost Center": ["IT", "HR"],
        "Planned": [10000, 12000],
        "Actual": [11000, 11500],
        "FX Rate": [1.0, 1.0]
    })

    buffer = _make_excel(df)

    response = client.post(
        "/analyze",
        files={"file": ("input.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert response.status_code == 200
    result = response.json()

    # ✅ API ออกมาเป็น string (formatted)
    assert result[0]["Planned"] == "10,000.00"
    assert result[0]["Actual"] == "11,000.00"
    assert result[0]["Variance"] == "1,000.00"
    assert result[1]["Planned"] == "12,000.00"
    assert result[1]["Actual"] == "11,500.00"
    assert result[1]["Variance"] == "-500.00"


def test_analyze_with_fx_rate_applied():
    df = pd.DataFrame({
        "Version": ["V1"],
        "Scenario": ["FX"],
        "Cost Center": ["Sales"],
        "Planned": [10000],
        "Actual": [10000],
        "FX Rate": [1.2],
    })

    buffer = _make_excel(df)

    response = client.post(
        "/analyze",
        files={"file": ("fx_input.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert response.status_code == 200
    r = response.json()[0]
    assert r["Planned"] == "10,000.00"
    assert r["Actual"] == "10,000.00"
    assert r["Variance"] == "2,000.00"  # 12,000 - 10,000


def test_download_report_returns_excel_and_values_are_numeric():
    df = pd.DataFrame({
        "Version": ["V1"],
        "Scenario": ["Base"],
        "Cost Center": ["Finance"],
        "Planned": [5000],
        "Actual": [5200]
    })

    buffer = _make_excel(df)

    response = client.post(
        "/download-report",
        files={"file": ("input.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert len(response.content) > 3000

    # โหลดกลับมาตรวจว่าเป็น numeric
    excel_buffer = BytesIO(response.content)
    wb = openpyxl.load_workbook(excel_buffer, data_only=True)
    ws = wb.active
    planned_value = ws["D2"].value
    actual_value = ws["E2"].value
    assert isinstance(planned_value, (int, float))
    assert isinstance(actual_value, (int, float))


def test_download_pdf_returns_pdf():
    df = pd.DataFrame({
        "Version": ["V1"],
        "Scenario": ["Base"],
        "Cost Center": ["Ops"],
        "Planned": [3000],
        "Actual": [2800]
    })

    buffer = _make_excel(df)

    response = client.post(
        "/download-pdf",
        files={"file": ("input.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert len(response.content) > 1500


def test_format_with_k_and_m_unit():
    assert number_format_utils.format_number(12_345, "k") == "12.35K"
    assert number_format_utils.format_number(2_500_000, "m") == "2.50M"
    assert number_format_utils.format_number(1234.56, "number") == "1,234.56"
