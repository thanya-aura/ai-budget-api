# budget_premium/modules/excel_dashboard.py
from io import BytesIO
from typing import List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from .number_format import apply_scale_series

MONEY_FMT = '#,##0.00'
PCT_FMT   = '0.00%'

# ใช้สเกลเดียวกับ output (raw/k/m) ทั้ง KPI และตาราง
SCALE_MAP = {"raw": 1, "k": 1_000, "m": 1_000_000}
def _scaled(value: float | None, scale: str) -> float | None:
    if value is None:
        return None
    factor = SCALE_MAP.get(scale.lower(), 1)
    return float(value) / factor

def _write_kpi(ws, row: int, name: str, value: float | None) -> int:
    ws[f"A{row}"] = name
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"] = float(value) if value is not None else None
    ws[f"B{row}"].number_format = MONEY_FMT
    return row + 1

def generate_excel_dashboard(df: pd.DataFrame, scale: str = "raw") -> BytesIO:
    """
    Executive dashboard (Excel):
      - KPI ย่อสเกลตาม 'scale' ให้ตรงกับตาราง
      - ตารางเงินฟอร์แมต #,##0.00 และเติมหน่วย (K/M) ที่หัวคอลัมน์เงิน
      - ตำแหน่ง header คุมด้วย ws.max_row+1 ป้องกันหัวซ้ำ
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Dashboard"

    # Title + scale indicator
    ws["A1"] = "Executive Budget Dashboard"
    ws["A1"].font = Font(bold=True, size=14)
    ws["F1"] = f"Scale: {scale.upper()}"

    # ---------- KPIs (ใช้สเกลเดียวกับตาราง) ----------
    tot_planned  = df["Planned"].sum()          if "Planned" in df.columns else None
    tot_actual   = df["Actual"].sum()           if "Actual" in df.columns else None
    tot_adjusted = df["Adjusted Actual"].sum()  if "Adjusted Actual" in df.columns else None
    tot_var      = df["Variance"].sum()         if "Variance" in df.columns else None

    row = 3
    if tot_planned  is not None: row = _write_kpi(ws, row, "Total Planned",  _scaled(tot_planned,  scale))
    if tot_actual   is not None: row = _write_kpi(ws, row, "Total Actual",   _scaled(tot_actual,   scale))
    if tot_adjusted is not None: row = _write_kpi(ws, row, "Total Adjusted", _scaled(tot_adjusted, scale))
    if tot_var      is not None: row = _write_kpi(ws, row, "Total Variance", _scaled(tot_var,      scale))

    # ---------- ตารางหลัก ----------
    table_cols: List[str] = [c for c in ["Cost Center","Project","Planned","Actual","Adjusted Actual","Variance"]
                             if c in df.columns]
    tdf = df[table_cols].copy()

    # ย่อสเกล "เฉพาะแสดงผล" ในคอลัมน์เงิน
    money_cols = [c for c in ["Planned","Actual","Adjusted Actual","Variance"] if c in tdf.columns]
    for c in money_cols:
        tdf[c] = apply_scale_series(tdf[c], scale=scale)

    # จับตำแหน่ง header ก่อน append เพื่อไม่ให้หัวซ้ำ
    header_row = ws.max_row + 1
    for r in dataframe_to_rows(tdf, index=False, header=True):
        ws.append(list(r) if not isinstance(r, list) else r)

    data_first = header_row + 1
    data_last  = ws.max_row
    col_count  = len(table_cols)

    thin = Side(style="thin", color="DDDDDD")
    name_to_idx = {name: idx for idx, name in enumerate(table_cols, start=1)}
    unit_suffix = {"k": " (K)", "m": " (M)"}.get(scale.lower(), "")

    # สไตล์หัว + เติมหน่วยที่หัวคอลัมน์เงิน
    for col_name, col_idx in name_to_idx.items():
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = col_name + (unit_suffix if col_name in money_cols else "")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # สไตล์บอดี้ + number format
    for r in range(data_first, data_last + 1):
        for c in range(1, col_count + 1):
            ws.cell(row=r, column=c).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for mcol in money_cols:
            cidx = name_to_idx[mcol]
            ws.cell(row=r, column=cidx).number_format = MONEY_FMT

    # ความกว้างคอลัมน์
    widths = {"Cost Center":18,"Project":24,"Planned":14,"Actual":14,"Adjusted Actual":16,"Variance":14}
    for idx, name in enumerate(table_cols, start=1):
        ws.column_dimensions[chr(64+idx)].width = widths.get(name, 14)

    bio = BytesIO()
    wb.save(bio); bio.seek(0)
    return bio
